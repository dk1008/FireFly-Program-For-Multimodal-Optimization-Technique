

				#################################################################################
				# 										#
				#			K_Means Clustering Algorithm In Python			#
				#										#
				#################################################################################


#*******************************************************************************************************************************************#
'''***************************************************************************************************************************************'''
#*******************************************************************************************************************************************#

import matplotlib.pyplot as plt
import openpyxl
import random
import math
print(" ")
print("Example of Path of Excel File: /home/lenovo/Documents/k-means_Clustering/glass_Data_set.xlsx")
print(" ")
Location = input("Enter Path Of Excel File: ")
wb = openpyxl.load_workbook(Location) 
sheet = wb.active




###################
## It Contains Data Set Values
NData = sheet.max_row
###################





###################
## It Take Dimension Of Data Set As Input
DD = sheet.max_column
###################





###################
## It Take Number Of Centroid As Input
NC = int(input("A Firefly Has Number Of Centriod: "))
###################





#START****************************************************************************************************************************************
# LB -> Lower Bound 	UB -> Upper Bound
LB = []	# It Take Lower Bound As Input
UB = []	# It take Upper Bound As Input
for i in range(DD):
	for j in range(NData):
		if(j == 0):
			LB.append(sheet.cell(row = j+1, column = i+1).value)
			UB.append(sheet.cell(row = j+1, column = i+1).value)
		else:
			num = sheet.cell(row = j+1, column = i+1).value
			LB[i] = min(LB[i] , num)
			UB[i] = max(UB[i] , num)
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START****************************************************************************************************************************************

MD = 0	#MD -> Max Distance Of Centroid DataSets 

# This Will Generate A Distance Which Will Be longest Distance Between any Points in centroid Data set
for i in range(DD):
	MD += math.pow((UB[i]-LB[i]),2)
MD = math.sqrt(MD)
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

#Distance Between Centroid And A Data Set Point
def Distance_Evaluation(CP , DP):	#CP -> Centroid Data Set	#DP -> Data Point DataSet
	sum = 0
	for i in range(DD):
		sum += math.pow((CP[i]-DP[i]),2)
	return (math.sqrt(abs(sum))) 	#return square of distance between two firefly..

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START****************************************************************************************************************************************

DPP = []	#DPP -> Input Data

#Generating Data Set
#DPP Contain "NData" Unit Of Data Set Value
#Each Data Set Has "DD" Unit Dimension
def Data_Point_Position(Location, wb, sheet, NData , DD):
	for i in range (NData):
		DPP.append([])
		for j in range (DD):
			DPP[i].append([])
			DPP[i][j] = sheet.cell(row = i+1, column = j+1).value #initializing the Input Data

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

Centroid = []	#Centroid -> List Of Centroid

#generating Fireflies
#Each Firefly has "NC" Unit of Centroid
#Each Centroid has "DD" Unit of Dimension
def Centroid_List(NC , DD , LB , UB):
	for i in range(NC):
		Centroid.append([])
		for j in range(DD):
			Centroid[i].append(random.uniform(LB[j] , UB[j]))

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

CRD = []	#CDL -> Centroid With Respective DataSet
def respective_Centroid_List(NC):
	CRD [:]=[]
	for i in range(NC):
		CRD.append([])

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

#Seprating Dataset According to Their Respective Centroid
#MD -> Max Distance Of Centroid Data Sets
#NData -> Number Of DataSet
#NC -> Each Firefly Has Number of centroid 
def DataSet_Centroid_Connection(MD , NData , NC ):
	for i in range(NData):
		CT_Index = 0	#Allocate A DataSet To A Respective Centroid
		dis = MD
		for j in range(NC):
			distance = Distance_Evaluation(DPP[i] , Centroid[j])
			if(distance <= dis):
				dis = distance
				CT_Index = j
		CRD[CT_Index].append(DPP[i])
	 
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

#Calculating and Allocating Dataset-Centroid Distance According to Their Respective Centroid
#MD -> Max Distance Of Centroid Data Sets
#NF -> Number Of Firefly
#NData -> Number Of DataSet
#NC -> Each Firefly Has Number of centroid 
def DataSet_Centroid_Distance_Connection(MD , NData , NC ):
	total = 0
	for i in range(NData):
		dis = MD
		for j in range(NC):
			distance = Distance_Evaluation(DPP[i] , Centroid[j])
			if(distance <= dis):
				dis = distance
		'''DBCD.append(dis) '''
		total += dis
	return total

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************
Error_Value = []
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

def Centroid_Average(NC , DD):
	for i in range(NC):
		for j in range(DD):
			total = 0
			avg = 0
			for k in range(len(CRD[i])):
				total += CRD[i][k][j]
			if(total != 0):
				avg = total/len(CRD[i])
			Centroid[i][j] = avg

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************
def Display(NC , DD):
	print(" ")
	for i in range(NC):
		print("Centroid-{0} is: ".format(i+1),end = " "),
		for j in range(DD):
			print("%.6f  "%(Centroid[i][j]), end = " "),
		print(" ")
			
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START*************************************************************************************************************************************
def Cluster_Ploting(NC): #here we have to plot all input data and centroid in different color..
	#plt.axis([math.floor(LB[0]),math.ceil(UB[0]),math.floor(LB[1]),math.ceil(UB[1])])
	for i in range(NData):
		plt.plot(DPP[i][0],DPP[i][1],"r o")
	for i in range(NC):
		for j in range(len(CRD[i])):
			if(len(CRD[i])>0):
				x = [Centroid[i][0],CRD[i][j][0]]
				y = [Centroid[i][1],CRD[i][j][1]]
				plt.plot(x,y,"y-")

	for i in range(NC):
		plt.plot(Centroid[i][0] , Centroid[i][1],"g *")
	plt.show()
		
'''END----------------------------------------------------------------------------------------------------------------------------------'''





#START*************************************************************************************************************************************
def Error_Ploting():
	for i in range(len(Error_Value)):
		plt.plot(i,Error_Value[i],"r o")
	plt.show()
'''END----------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************
test = []
def Test_Centroid(NC , DD):
	for i in range(NC):
		test.append([])
		for j in range(DD):
			test[i].append(0)

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************
#Siloutee Value plotting...
def SI_Ploting():
	for i in range(len(SI)):
		plt.plot(i,SI[i],"g o")
	plt.show()

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START*************************************************************************************************************************************
IntraCD = []
for i in range(NC):
	IntraCD.append([])
def Intra_Cluster_distance(NC):
	for i in range(NC):
		for j in range(len(CRD[i])):
			total = 0
			if(len(CRD[i]) > 0):
				for k in range(len(CRD[i])):
					dis = Distance_Evaluation(CRD[i][j] , CRD[i][k])
					total += dis
			IntraCD[i].append(total)
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START**********************************************************************************************************************************
InterCD = []
for i in range(NC):
	InterCD.append([])
def Inter_Cluster_distance(NC):
	for i in range(NC):
		for j in range(len(CRD[i])):
			total = 0
			if(len(CRD[i]) > 0):
				for k in range(NC):
					if(i != k):
						for l in range(len(CRD[k])):
							if(len(CRD[k]) > 0):
								dis = Distance_Evaluation(CRD[i][j] , CRD[k][l])
								total += dis
				InterCD[i].append(total)
'''END---------------------------------------------------------------------------------------------------------------------------------'''





#START***********************************************************************************************************************************
SI = []
def siloutte_Indexing(NC):
	total = 0
	x = 0
	Intra_Cluster_distance(NC)
	Inter_Cluster_distance(NC)
	for i in range(NC):
		for j in range(len(IntraCD[i])):
			if(len(IntraCD[i]) > 0):
				x = (InterCD[i][j] - IntraCD[i][j])/max(IntraCD[i][j],InterCD[i][j])
		total += x
		total = total/NC
	SI.append(total) 
'''END---------------------------------------------------------------------------------------------------------------------------------'''





#START****************************************************************************************************************************************
def saving_Error_in_excel():
	result_location = ("/home/lenovo/Documents/K_Mean_Firefly_clustering/RESULT_FE_K.xlsx")
	wx = openpyxl.load_workbook(result_location)
	sheet1 = wx.active
	no_of_row = sheet1.max_row
	c1 = sheet1.cell(row = no_of_row+1, column = 4)
	c1.value = Error_Value[len(Error_Value)-1]
	wx.save("/home/lenovo/Documents/K_Mean_Firefly_clustering/RESULT_FE_K.xlsx")
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START****************************************************************************************************************************************
def saving_file_in_excel():
	result_location = ("/home/lenovo/Documents/K_Mean_Firefly_clustering/RESULT_F_K.xlsx")
	ws = openpyxl.load_workbook(result_location)
	sheet1 = ws.active
	no_of_row = sheet1.max_row
	c1 = sheet1.cell(row = no_of_row+1, column = 4)
	c1.value = SI[len(SI)-1]
	ws.save("/home/lenovo/Documents/K_Mean_Firefly_clustering/RESULT_F_K.xlsx")
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************
#K-Mean Algorithm.....

Data_Point_Position(Location, wb, sheet, NData , DD)
Centroid_List(NC , DD , LB , UB)
Test_Centroid(NC , DD)
DBCD = DataSet_Centroid_Distance_Connection(MD ,NData , NC )
Error_Value.append(DBCD)

Display(NC , DD)
loop=0
while(True):
	respective_Centroid_List(NC)
	DataSet_Centroid_Connection(MD , NData , NC )
	Centroid_Average(NC , DD)
	DBCD = DataSet_Centroid_Distance_Connection(MD ,NData , NC )
	Error_Value.append(DBCD)
	siloutte_Indexing(NC)
	Display(NC, DD)
	if(DD == 2):
		Cluster_Ploting(NC)
	flag = 0
	for i in range(NC):
		for j in range(DD):
			if(Centroid[i][j] != test[i][j]):
				test[i][j] = Centroid[i][j]
				flag = 1
	if(flag == 0):
		if(loop <=5):
			loop += 1
			flag = 1
		else:
			break	

saving_Error_in_excel()
saving_file_in_excel()
Error_Ploting()	#Error value displaying...
SI_Ploting()  # Siloutee value displaying.
'''
#print("MAX DISTANCE ", MD)
#print()
print("Input Data is")
print(DPP)
print()
print("Centroid with respective data")
print(CRD)
print()
print("radius is ")
print(radius)'''
'''END------------------------------------------------------------------------------------------------------------------------------------'''

